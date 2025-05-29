import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import os
import shutil
import re
import ctypes
import math

# 设置高 DPI 缩放
ctypes.windll.shcore.SetProcessDpiAwareness(1)

def browse_file():
    """浏览文件并设置文件路径"""
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        entry_file_path.delete(0, tk.END)
        entry_file_path.insert(0, file_path)
        load_sheets(file_path)

def load_sheets(file_path):
    """加载Excel文件中的所有Sheet名称"""
    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        combo_sheet['values'] = sheets
        combo_sheet.current(0)  # 默认选择第一个Sheet
        load_columns(file_path, sheets[0])  # 加载第一个Sheet的列名
        wb.close()
    except Exception as e:
        show_error_message("错误", f"无法读取文件: {e}")

def load_columns(file_path, sheet_name):
    """加载指定Sheet的列名到下拉框"""
    try:
        wb = load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        
        # 获取第一行的所有值作为列名
        columns = [cell.value for cell in ws[1] if cell.value is not None]
        combo_column['values'] = columns
        combo_column.current(0)  # 默认选择第一列
        wb.close()
    except Exception as e:
        show_error_message("错误", f"无法读取Sheet: {e}")

def on_sheet_select(event):
    """当用户选择Sheet时，加载该Sheet的列名"""
    file_path = entry_file_path.get()
    if file_path:
        selected_sheet = combo_sheet.get()
        load_columns(file_path, selected_sheet)

def clean_filename(filename):
    """清理文件名中的非法字符"""
    if not isinstance(filename, str):
        filename = str(filename)  # 确保传入的是字符串
    illegal_chars = r'[\\/:*?"<>|]'
    return re.sub(illegal_chars, '_', filename)

def auto_adjust_column_widths(ws, max_rows=100):
    """自动调整Excel列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # 计算最大长度，最多检查前100行以避免性能问题
        for cell in column[:max_rows]:
            try:
                value = str(cell.value) if cell.value is not None else ""
                if len(value) > max_length:
                    max_length = len(value)
            except:
                pass
        
        adjusted_width = (max_length + 2) * 1.5  # 加一点缓冲空间
        ws.column_dimensions[column_letter].width = adjusted_width

def get_column_values(ws, column_name, has_header=True):
    """获取指定列的所有值"""
    column_index = None
    header_row = 1 if has_header else 0
    
    # 查找列名所在的列索引
    for cell in ws[1 if has_header else 0]:
        if cell.value == column_name:
            column_index = cell.column
            break
    
    if column_index is None:
        return []
    
    # 获取该列的所有值
    values = []
    for row in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index)
        values.append(cell.value)
    
    return values

def get_group_data(file_path, sheet_name, column_name):
    """获取按指定列分组的数据"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    # 获取列索引
    column_index = None
    for cell in ws[1]:  # 假设第一行是标题行
        if cell.value == column_name:
            column_index = cell.column
            break
    
    if column_index is None:
        wb.close()
        return {}
    
    # 收集所有唯一值
    unique_values = set()
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=column_index)
        if cell.value is not None:
            unique_values.add(cell.value)
    
    # 为每个唯一值收集行数据
    grouped_data = {}
    for value in unique_values:
        grouped_data[value] = []
    
    # 收集每行数据
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            row_data.append(cell.value)
        
        cell_value = ws.cell(row=row, column=column_index).value
        if cell_value in grouped_data:
            grouped_data[cell_value].append(row_data)
    
    wb.close()
    return grouped_data

def split_excel():
    """执行拆分操作"""
    file_path = entry_file_path.get()
    sheet_name = combo_sheet.get()
    column_name = combo_column.get()

    if not file_path or not sheet_name or not column_name:
        show_warning_message("警告", "请选择文件、Sheet和拆分列！")
        return

    try:
        grouped_data = get_group_data(file_path, sheet_name, column_name)
        if not grouped_data:
            show_warning_message("警告", "没有找到可拆分的数据！")
            return

        progress_bar['maximum'] = len(grouped_data)
        progress_bar['value'] = 0
        root.update_idletasks()

        # 获取原始文件名（不带扩展名）
        base_filename = os.path.splitext(os.path.basename(file_path))[0]
        
        # 创建输出文件夹（使用源文件所在目录）
        source_dir = os.path.dirname(file_path)
        output_folder = os.path.join(source_dir, f"{base_filename}-拆分后")
        if os.path.exists(output_folder):
            shutil.rmtree(output_folder)  # 如果文件夹已存在，先删除
        os.makedirs(output_folder)

        if var_split_to_files.get():
            # 拆分为单独文件
            wb_template = load_workbook(file_path, read_only=True)
            template_ws = wb_template[sheet_name]
            
            for name, rows in grouped_data.items():
                clean_name = clean_filename(name)
                output_file = os.path.join(output_folder, f'{base_filename}-{clean_name}.xlsx')
                wb = Workbook()
                wb.remove(wb.active)  # 删除默认创建的工作表
                
                # 复制模板工作表
                new_ws = wb.create_sheet(title=sheet_name)
                
                # 复制标题行
                for col, cell in enumerate(template_ws[1], 1):
                    new_cell = new_ws.cell(row=1, column=col, value=cell.value)
                    # 如果是超长数字（如身份证号），则设置为文本格式
                    if isinstance(cell.value, (int, float)) and len(str(int(cell.value))) > 15:
                        new_cell.number_format = '@'
                
                # 复制数据行
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = new_ws.cell(row=row_idx, column=col_idx, value=value)
                        # 如果是超长数字（如身份证号），则设置为文本格式
                        if isinstance(value, (int, float)) and len(str(int(value))) > 15:
                            cell.number_format = '@'
                
                # 自动调整列宽
                auto_adjust_column_widths(new_ws)
                wb.save(output_file)
                print(f'Saved {output_file}')
                progress_bar['value'] += 1
                root.update_idletasks()
            
            wb_template.close()
        else:
            # 拆分到同一文件的不同工作表
            output_file = os.path.join(output_folder, f'{base_filename}-拆分后.xlsx')
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认创建的工作表
            
            # 读取模板工作表的第一行作为标题
            wb_template = load_workbook(file_path, read_only=True)
            template_ws = wb_template[sheet_name]
            headers = [cell.value for cell in template_ws[1] if cell.value is not None]
            wb_template.close()
            
            for name, rows in grouped_data.items():
                clean_name = clean_filename(name)
                ws = wb.create_sheet(title=clean_name[:31])  # Excel sheet名称最多31个字符
                
                # 写入标题行
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col_idx, value=header)
                
                # 写入数据行
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        # 如果是超长数字（如身份证号），则设置为文本格式
                        if isinstance(value, (int, float)) and len(str(int(value))) > 15:
                            cell.number_format = '@'
                
                # 自动调整列宽
                auto_adjust_column_widths(ws)
                progress_bar['value'] += 1
                root.update_idletasks()
            
            wb.save(output_file)
            print(f'Saved {output_file}')

        show_info_message("完成", f"拆分完成！结果已保存到文件夹: {output_folder}")
    except Exception as e:
        show_error_message("错误", f"拆分失败: {e}")
    finally:
        progress_bar['value'] = 0

def show_info_message(title, message):
    """在窗口中心显示信息消息"""
    show_message(title, message, "info")

def show_warning_message(title, message):
    """在窗口中心显示警告消息"""
    show_message(title, message, "warning")

def show_error_message(title, message):
    """在窗口中心显示错误消息"""
    show_message(title, message, "error")

def show_message(title, message, icon_type):
    """在窗口中心显示自定义消息"""
    top = tk.Toplevel(root)
    top.title(title)
    top.geometry(f"+{root.winfo_x() + (root.winfo_width() // 2) - 100}+{root.winfo_y() + (root.winfo_height() // 2) - 50}")
    label = tk.Label(top, text=message, padx=10, pady=10)
    label.pack()
    if icon_type == "info":
        button = tk.Button(top, text="确定", command=top.destroy)
    elif icon_type == "warning":
        button = tk.Button(top, text="确定", command=top.destroy)
    elif icon_type == "error":
        button = tk.Button(top, text="确定", command=top.destroy)
    button.pack(pady=5)
    top.transient(root)  # 使Toplevel窗口总是位于父窗口上方
    top.grab_set()      # 设置输入焦点到此窗口
    root.wait_window(top)

# 创建主窗口
root = tk.Tk()
root.title("Excel拆分工具")
font_style = ("微软雅黑", 10)
root.option_add("*Font", font_style)

label_file_path = tk.Label(root, text="选择Excel文件:")
label_file_path.grid(row=0, column=0, padx=5, pady=5, sticky="e")

entry_file_path = tk.Entry(root, width=50)
entry_file_path.grid(row=0, column=1, padx=5, pady=5)

button_browse = tk.Button(root, text="浏览", command=browse_file)
button_browse.grid(row=0, column=2, padx=5, pady=5)

# Sheet选择
label_sheet = tk.Label(root, text="选择Sheet:")
label_sheet.grid(row=1, column=0, padx=5, pady=5, sticky="e")  # 右对齐

combo_sheet = ttk.Combobox(root, width=47)
combo_sheet.grid(row=1, column=1, padx=5, pady=5)
combo_sheet.bind("<<ComboboxSelected>>", on_sheet_select)

# 列选择
label_column = tk.Label(root, text="选择拆分列:")
label_column.grid(row=2, column=0, padx=5, pady=5, sticky="e")  # 右对齐

combo_column = ttk.Combobox(root, width=47)
combo_column.grid(row=2, column=1, padx=5, pady=5)

# 进度条
label_progress = tk.Label(root, text="进度:")
label_progress.grid(row=3, column=0, padx=5, pady=10, sticky="e")  # 右对齐

progress_bar = ttk.Progressbar(root, orient="horizontal", length=350, mode="determinate")
progress_bar.grid(row=3, column=1, padx=5, pady=10, sticky="ew")  # 左右对齐

# 拆分选项
var_split_to_files = tk.BooleanVar(value=True)  # 默认勾选
check_split_to_files = tk.Checkbutton(root, text="拆分为单独文件", variable=var_split_to_files)
check_split_to_files.grid(row=4, column=0, padx=5, pady=10, sticky="w")  # 左对齐

# 拆分按钮
button_split = tk.Button(root, text="拆分", command=split_excel)
button_split.grid(row=4, column=1, padx=5, pady=10)

# 运行主循环
root.mainloop()
